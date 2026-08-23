from __future__ import annotations

import hashlib
import math
import re
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

SUPPORTED_VIEWS = ("side", "back")
SOURCE_DATASET = "mendeley_horqin_side_back_v3"
SOURCE_DOI = "10.17632/h2s22wr5py.3"
MIN_EXPECTED_SHORT_EDGE_PX = 2000
KNOWN_V3_DUPLICATE_BACK_IDS = ("57", "67")

_MEASUREMENT_COLUMNS = {
    "num": "source_animal_id",
    "oblique body length (cm)": "oblique_body_length_cm",
    "withers height(cm)": "withers_height_cm",
    "heart girth(cm)": "heart_girth_cm",
    "hip length (cm)": "hip_length_cm",
    "body weight (kg)": "weight_kg",
}
_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png"}


def resolve_horqin_root(dataset_root: str | Path) -> Path:
    """Resolve a pasta que contém measurements.xlsx, side view e back view."""
    root = Path(dataset_root)
    if not root.is_dir():
        raise FileNotFoundError(f"Diretório do dataset não encontrado: {root}")
    if (root / "measurements.xlsx").is_file():
        return root
    candidates = [path.parent for path in root.rglob("measurements.xlsx")]
    if len(candidates) != 1:
        raise ValueError(
            "Era esperada exatamente uma pasta com measurements.xlsx; "
            f"foram encontradas {len(candidates)}."
        )
    return candidates[0]


def read_horqin_measurements(
    dataset_root: str | Path,
) -> dict[str, dict[str, str]]:
    """Lê pesos e medidas corporais do arquivo original."""
    root = resolve_horqin_root(dataset_root)
    workbook = load_workbook(
        root / "measurements.xlsx", read_only=True, data_only=True
    )
    sheet = workbook.active
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [
        _MEASUREMENT_COLUMNS.get(_normalize_header(value)) for value in raw_headers
    ]
    if not {"source_animal_id", "weight_kg"}.issubset(set(headers)):
        workbook.close()
        raise ValueError("measurements.xlsx não contém identificação e peso.")

    measurements: dict[str, dict[str, str]] = {}
    for line_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        row = {
            header: _format_cell(value)
            for header, value in zip(headers, values, strict=False)
            if header and value is not None
        }
        if not row:
            continue
        source_id = row.get("source_animal_id", "")
        if not source_id.isdigit() or int(source_id) <= 0:
            workbook.close()
            raise ValueError(f"linha {line_number}: Num inválido: {source_id!r}")
        source_id = str(int(source_id))
        if source_id in measurements:
            workbook.close()
            raise ValueError(f"Animal duplicado na planilha: {source_id}")
        try:
            weight = float(row.get("weight_kg", ""))
        except ValueError as exc:
            workbook.close()
            raise ValueError(f"Peso inválido para animal {source_id}.") from exc
        if not math.isfinite(weight) or weight <= 0:
            workbook.close()
            raise ValueError(f"Peso inválido para animal {source_id}.")
        row["source_animal_id"] = source_id
        measurements[source_id] = row
    workbook.close()
    if not measurements:
        raise ValueError("Nenhuma medida foi encontrada.")
    return measurements


def scan_horqin_images(dataset_root: str | Path) -> dict[str, dict[str, Path]]:
    """Indexa as duas vistas por animal, aceitando PNG e o JPG publicado."""
    root = resolve_horqin_root(dataset_root)
    images: dict[str, dict[str, Path]] = {}
    for view, directory_name in (("side", "side view"), ("back", "back view")):
        directory = root / directory_name
        if not directory.is_dir():
            raise FileNotFoundError(f"Pasta de vista ausente: {directory}")
        for path in sorted(directory.iterdir()):
            if not path.is_file() or path.suffix.lower() not in _IMAGE_SUFFIXES:
                continue
            if not path.stem.isdigit() or int(path.stem) <= 0:
                raise ValueError(f"Nome de imagem inesperado: {path.name}")
            source_id = str(int(path.stem))
            animal_images = images.setdefault(source_id, {})
            if view in animal_images:
                raise ValueError(
                    f"Imagem duplicada para animal {source_id}, vista {view}."
                )
            animal_images[view] = path
    if not images:
        raise ValueError("Nenhuma imagem Horqin foi encontrada.")
    return images


def find_horqin_anomalies(
    dataset_root: str | Path,
    *,
    views: Iterable[str] = SUPPORTED_VIEWS,
    min_short_edge_px: int = MIN_EXPECTED_SHORT_EDGE_PX,
) -> dict[str, tuple[str, ...]]:
    """Lista, por animal, vistas ausentes ou com resolução fora do corpus."""
    selected_views = _validate_views(views)
    measurements = read_horqin_measurements(dataset_root)
    images = scan_horqin_images(dataset_root)
    anomalies: dict[str, tuple[str, ...]] = {}
    for source_id in sorted(measurements, key=int):
        reasons: list[str] = []
        for view in selected_views:
            path = images.get(source_id, {}).get(view)
            if path is None:
                reasons.append(f"missing_{view}")
                continue
            with Image.open(path) as image:
                width, height = image.size
            if min(width, height) < min_short_edge_px:
                reasons.append(f"low_resolution_{view}:{width}x{height}")
        if reasons:
            anomalies[source_id] = tuple(reasons)
    if "back" in selected_views:
        duplicate_paths = [
            images.get(source_id, {}).get("back")
            for source_id in KNOWN_V3_DUPLICATE_BACK_IDS
        ]
        if all(path is not None for path in duplicate_paths) and len(
            {_sha256(path) for path in duplicate_paths if path is not None}
        ) == 1:
            for source_id in KNOWN_V3_DUPLICATE_BACK_IDS:
                reasons = list(anomalies.get(source_id, ()))
                reasons.append("exact_duplicate_back_across_animals:57,67")
                anomalies[source_id] = tuple(reasons)
    return anomalies


def build_horqin_manifest_rows(
    dataset_root: str | Path,
    *,
    image_root: str | Path,
    views: Iterable[str] = ("side",),
    exclude_known_anomalies: bool = False,
    min_short_edge_px: int = MIN_EXPECTED_SHORT_EDGE_PX,
) -> list[dict[str, str]]:
    """Compõe o manifesto Horqin, exigindo decisão explícita sobre anomalias."""
    selected_views = _validate_views(views)
    root = resolve_horqin_root(dataset_root)
    measurements = read_horqin_measurements(root)
    images = scan_horqin_images(root)
    image_base = Path(image_root).resolve()

    extra_images = sorted(set(images) - set(measurements), key=int)
    if extra_images:
        raise ValueError("Imagens sem medidas: " + ", ".join(extra_images))
    anomalies = find_horqin_anomalies(
        root,
        views=selected_views,
        min_short_edge_px=min_short_edge_px,
    )
    if anomalies and not exclude_known_anomalies:
        details = "; ".join(
            f"{animal}: {', '.join(reasons)}"
            for animal, reasons in list(anomalies.items())[:10]
        )
        raise ValueError(
            "Anomalias de imagem exigem exclusão explícita com "
            f"exclude_known_anomalies=True: {details}"
        )

    rows: list[dict[str, str]] = []
    for source_id in sorted(measurements, key=int):
        if source_id in anomalies:
            continue
        measurement = measurements[source_id]
        animal_id = f"horqin_{int(source_id):03d}"
        for view in selected_views:
            image_path = images[source_id][view]
            relative_path = _relative_to_image_root(image_path, image_base)
            rows.append(
                {
                    "image_path": relative_path.as_posix(),
                    "animal_id": animal_id,
                    "event_id": f"{animal_id}_capture_001",
                    "weight_kg": measurement["weight_kg"],
                    "view": view,
                    "breed": "horqin",
                    "environment": "grassland",
                    "quality": "source_original",
                    "label_status": "published_scale_measurement",
                    "training_eligible": "research_only_until_domain_validation",
                    "source_dataset": SOURCE_DATASET,
                    "source_version": "3",
                    "source_doi": SOURCE_DOI,
                    "source_license": "CC_BY_4_0",
                    "source_animal_id": source_id,
                    **{
                        key: value
                        for key, value in measurement.items()
                        if key not in {"source_animal_id", "weight_kg"}
                    },
                }
            )
    if not rows:
        raise ValueError("Nenhuma amostra restou após aplicar as exclusões.")
    return rows


def _validate_views(views: Iterable[str]) -> tuple[str, ...]:
    selected = tuple(dict.fromkeys(views))
    invalid = set(selected) - set(SUPPORTED_VIEWS)
    if invalid:
        raise ValueError(f"Vistas Horqin inválidas: {sorted(invalid)}")
    if not selected:
        raise ValueError("Selecione ao menos uma vista Horqin.")
    return selected


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _format_cell(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _relative_to_image_root(path: Path, image_root: Path) -> Path:
    try:
        return path.resolve().relative_to(image_root)
    except ValueError as exc:
        raise ValueError(
            f"Imagem {path} não está dentro de image_root {image_root}."
        ) from exc


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()

from __future__ import annotations

import re
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

SUPPORTED_VIEWS = ("left", "right", "top")

MEASUREMENT_COLUMNS = {
    "n": "source_animal_id",
    "live weithg": "weight_kg",  # grafia presente no arquivo original
    "live weight": "weight_kg",
    "withers height": "withers_height_cm",
    "hip height": "hip_height_cm",
    "chest depth": "chest_depth_cm",
    "chest width": "chest_width_cm",
    "ilium width": "ilium_width_cm",
    "hip joint width": "hip_joint_width_cm",
    "oblique body length": "oblique_body_length_cm",
    "hip length": "hip_length_cm",
    "heart girth": "heart_girth_cm",
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", text)


def read_cowdb_measurements(path: str | Path) -> dict[str, dict[str, str]]:
    """Lê exclusivamente a planilha de medidas original do CowDB."""
    spreadsheet_path = Path(path)
    if not spreadsheet_path.is_file():
        raise FileNotFoundError(f"Planilha do CowDB não encontrada: {spreadsheet_path}")

    workbook = load_workbook(spreadsheet_path, read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapped_headers = [
        MEASUREMENT_COLUMNS.get(_normalize_header(header)) for header in raw_headers
    ]
    required = {"source_animal_id", "weight_kg"}
    if not required.issubset(set(mapped_headers)):
        raise ValueError(
            "A planilha CowDB não contém as colunas de identificação e peso esperadas."
        )

    measurements: dict[str, dict[str, str]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {
            header: _format_cell(value)
            for header, value in zip(mapped_headers, values, strict=False)
            if header and value is not None
        }
        source_id = row.get("source_animal_id", "")
        if not source_id:
            continue
        if source_id in measurements:
            raise ValueError(f"Animal duplicado na planilha CowDB: {source_id}")
        measurements[source_id] = row
    workbook.close()

    if not measurements:
        raise ValueError("Nenhuma medida foi encontrada na planilha CowDB.")
    return measurements


def _format_cell(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def scan_cowdb_rgb_images(
    dataset_root: str | Path,
    *,
    views: Iterable[str],
) -> dict[str, dict[str, Path]]:
    """Localiza exclusivamente as imagens RGB esperadas na árvore do CowDB."""
    root = Path(dataset_root)
    if not root.is_dir():
        raise FileNotFoundError(f"Diretório CowDB não encontrado: {root}")

    selected_views = tuple(dict.fromkeys(views))
    invalid_views = set(selected_views) - set(SUPPORTED_VIEWS)
    if invalid_views:
        raise ValueError(f"Vistas CowDB inválidas: {sorted(invalid_views)}")
    if not selected_views:
        raise ValueError("Selecione ao menos uma vista do CowDB.")

    images: dict[str, dict[str, Path]] = {}
    animal_directories = sorted(
        (path for path in root.iterdir() if path.is_dir() and path.name.isdigit()),
        key=lambda path: int(path.name),
    )
    for animal_directory in animal_directories:
        animal_images: dict[str, Path] = {}
        for view in selected_views:
            candidates = sorted((animal_directory / "raw" / view).glob("rgb-*.png"))
            if len(candidates) > 1:
                raise ValueError(
                    f"Mais de uma imagem RGB para animal {animal_directory.name}, "
                    f"vista {view}: {candidates}"
                )
            if candidates:
                animal_images[view] = candidates[0]
        missing_views = set(selected_views) - set(animal_images)
        if animal_images and missing_views:
            raise ValueError(
                f"Vistas RGB ausentes para animal {animal_directory.name}: "
                f"{sorted(missing_views)}"
            )
        if animal_images:
            images[animal_directory.name] = animal_images

    if not images:
        raise ValueError("Nenhuma imagem RGB CowDB foi encontrada.")
    return images


def build_cowdb_manifest_rows(
    dataset_root: str | Path,
    measurements_path: str | Path,
    *,
    image_root: str | Path,
    views: Iterable[str] = ("left",),
) -> list[dict[str, str]]:
    """Compõe medidas e imagens CowDB no contrato de manifesto MS-PESO."""
    measurements = read_cowdb_measurements(measurements_path)
    images = scan_cowdb_rgb_images(dataset_root, views=views)
    image_base = Path(image_root).resolve()

    missing_measurements = sorted(set(images) - set(measurements), key=int)
    if missing_measurements:
        raise ValueError(
            "Imagens sem medidas para os animais: " + ", ".join(missing_measurements)
        )
    missing_images = sorted(set(measurements) - set(images), key=int)
    if missing_images:
        raise ValueError(
            "Medidas sem imagens nas vistas selecionadas para os animais: "
            + ", ".join(missing_images)
        )

    rows: list[dict[str, str]] = []
    for source_id in sorted(measurements, key=int):
        measurement = measurements[source_id]
        animal_id = f"cowdb_{int(source_id):03d}"
        for view, image_path in sorted(images[source_id].items()):
            try:
                relative_image_path = image_path.resolve().relative_to(image_base)
            except ValueError as exc:
                raise ValueError(
                    f"Imagem {image_path} não está dentro de image_root {image_base}."
                ) from exc
            rows.append(
                {
                    "image_path": relative_image_path.as_posix(),
                    "animal_id": animal_id,
                    "event_id": f"{animal_id}_capture_001",
                    "weight_kg": measurement["weight_kg"],
                    "view": view,
                    "breed": "hereford",
                    "farm_id": "cowdb_farm",
                    "quality": "accepted",
                    "source_dataset": "CowDB",
                    "source_animal_id": source_id,
                    **{
                        key: value
                        for key, value in measurement.items()
                        if key not in {"source_animal_id", "weight_kg"}
                    },
                }
            )
    return rows
